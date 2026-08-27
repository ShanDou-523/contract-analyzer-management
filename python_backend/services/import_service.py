"""Parse, validate, and commit staged contract spreadsheet imports."""

from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.contract import Contract, ContractImportJob

MAX_IMPORT_ROWS = 5000
SUPPORTED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}
FIELD_ALIASES = {
    "contract_no": {"contract_no", "contractno", "合同编号", "编号"},
    "name": {"name", "合同名称", "名称", "合同名"},
    "category": {"category", "类别", "合同类别"},
    "party_a_name": {"party_a_name", "甲方", "甲方名称"},
    "party_b_name": {"party_b_name", "乙方", "乙方名称"},
    "project_name": {"project_name", "项目", "项目名称"},
    "department_name": {"department_name", "部门", "部门名称"},
    "sign_date": {"sign_date", "签署日期", "签订日期"},
    "effective_date": {"effective_date", "生效日期"},
    "start_date": {"start_date", "开始日期", "履约开始日期"},
    "end_date": {"end_date", "结束日期", "履约结束日期"},
    "amount": {"amount", "金额", "合同金额"},
    "currency": {"currency", "币种"},
    "tax_included": {"tax_included", "含税", "是否含税"},
    "risk_level": {"risk_level", "风险等级"},
    "status": {"status", "状态"},
}
ALLOWED_FIELDS = set(FIELD_ALIASES)
STATUS_VALUES = {"draft", "active", "expired", "terminated"}
RISK_VALUES = {"low", "medium", "high", "critical"}


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _canonical_header(value) -> str:
    header = _text(value).lower().replace(" ", "").replace("-", "_")
    for field, aliases in FIELD_ALIASES.items():
        if header in {alias.lower().replace(" ", "").replace("-", "_") for alias in aliases}:
            return field
    return header


def _normalize_row(row: dict) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        canonical = _canonical_header(key)
        if canonical in ALLOWED_FIELDS:
            normalized[canonical] = _text(value)
    return normalized


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("gb18030")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV 缺少表头")
    rows = [_normalize_row(row) for row in reader]
    return [_canonical_header(item) for item in reader.fieldnames], rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = []
        for values_row in sheet.iter_rows(values_only=True):
            values.append(values_row)
            if len(values) > MAX_IMPORT_ROWS + 1:
                break
    except Exception as exc:
        raise HTTPException(status_code=400, detail="XLSX 文件无法读取") from exc
    finally:
        try:
            workbook.close()
        except UnboundLocalError:
            pass
    if not values or not any(value is not None for value in values[0]):
        raise HTTPException(status_code=400, detail="XLSX 缺少表头")
    headers = [_canonical_header(value) for value in values[0]]
    rows = []
    for values_row in values[1:]:
        rows.append(_normalize_row(dict(zip(headers, values_row))))
    return headers, rows


def parse_import_file(filename: str, content: bytes) -> tuple[str, list[str], list[dict[str, str]]]:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_IMPORT_EXTENSIONS:
        raise HTTPException(status_code=400, detail="导入仅支持 .xlsx 或 .csv 文件")
    if extension == ".csv":
        columns, rows = _parse_csv(content)
    else:
        columns, rows = _parse_xlsx(content)
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=400, detail=f"单次导入最多支持 {MAX_IMPORT_ROWS} 行")
    return extension[1:], columns, rows


def _parse_date(value: str) -> date | None:
    if not value:
        return None
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(value, pattern).date()
        except ValueError:
            continue
    raise ValueError("日期必须是 YYYY-MM-DD")


def _parse_amount(value: str) -> Decimal | None:
    if not value:
        return None
    try:
        amount = Decimal(value.replace(",", "").replace("￥", "").replace("¥", ""))
    except InvalidOperation as exc:
        raise ValueError("金额格式无效") from exc
    if amount < 0:
        raise ValueError("金额不能为负数")
    if len(amount.as_tuple().digits) > 18 or -amount.as_tuple().exponent > 2:
        raise ValueError("金额最多18位且保留2位小数")
    return amount


def _parse_bool(value: str) -> bool | None:
    if not value:
        return None
    if value.lower() in {"1", "true", "yes", "y", "是", "含税"}:
        return True
    if value.lower() in {"0", "false", "no", "n", "否", "不含税"}:
        return False
    raise ValueError("含税字段必须是 是/否")


def normalize_contract_row(row: dict[str, str], row_number: int) -> tuple[dict, list[dict]]:
    errors: list[dict] = []
    name = row.get("name", "").strip()
    if not name:
        errors.append({"row": row_number, "field": "name", "message": "合同名称不能为空"})
    if len(name) > 512:
        errors.append({"row": row_number, "field": "name", "message": "合同名称不能超过512个字符"})
    parsed: dict = {field: row.get(field, "").strip() or None for field in ALLOWED_FIELDS}
    parsed["name"] = name
    try:
        parsed["sign_date"] = _parse_date(row.get("sign_date", ""))
        parsed["effective_date"] = _parse_date(row.get("effective_date", ""))
        parsed["start_date"] = _parse_date(row.get("start_date", ""))
        parsed["end_date"] = _parse_date(row.get("end_date", ""))
    except ValueError as exc:
        errors.append({"row": row_number, "field": "date", "message": str(exc)})
    try:
        parsed["amount"] = _parse_amount(row.get("amount", ""))
    except ValueError as exc:
        errors.append({"row": row_number, "field": "amount", "message": str(exc)})
    try:
        parsed["tax_included"] = _parse_bool(row.get("tax_included", ""))
    except ValueError as exc:
        errors.append({"row": row_number, "field": "tax_included", "message": str(exc)})
    currency = row.get("currency", "").strip().upper() or "CNY"
    if len(currency) != 3 or not currency.isascii() or not currency.isalpha():
        errors.append({"row": row_number, "field": "currency", "message": "币种必须是3位字母"})
    parsed["currency"] = currency
    risk_level = row.get("risk_level", "").strip().lower() or "medium"
    if risk_level not in RISK_VALUES:
        errors.append({"row": row_number, "field": "risk_level", "message": "风险等级无效"})
    parsed["risk_level"] = risk_level
    status = row.get("status", "").strip().lower() or "draft"
    if status not in STATUS_VALUES:
        errors.append({"row": row_number, "field": "status", "message": "合同状态无效"})
    parsed["status"] = status
    if (
        isinstance(parsed["start_date"], date)
        and isinstance(parsed["end_date"], date)
        and parsed["start_date"] > parsed["end_date"]
    ):
        errors.append({"row": row_number, "field": "date", "message": "履约开始日期不能晚于结束日期"})
    if row.get("contract_no") and len(row["contract_no"].strip()) > 128:
        errors.append({"row": row_number, "field": "contract_no", "message": "合同编号不能超过128个字符"})
    return parsed, errors


def validate_rows(db: Session, job: ContractImportJob) -> dict:
    rows = json.loads(job.rows_json or "[]")
    errors: list[dict] = []
    seen_numbers: set[str] = set()
    if not rows:
        return {"valid": False, "errors": [{"row": 0, "field": "file", "message": "导入文件没有数据行"}], "valid_rows": 0}
    for index, row in enumerate(rows, start=2):
        parsed, row_errors = normalize_contract_row(row, index)
        contract_no = (row.get("contract_no") or "").strip()
        if contract_no:
            if contract_no in seen_numbers:
                row_errors.append({"row": index, "field": "contract_no", "message": "文件内合同编号重复"})
            seen_numbers.add(contract_no)
            if (
                db.query(Contract)
                .filter(
                    Contract.organization_id == job.organization_id,
                    Contract.contract_no == contract_no,
                )
                .first()
            ):
                row_errors.append({"row": index, "field": "contract_no", "message": "合同编号已存在，导入不会覆盖"})
        errors.extend(row_errors)
    return {"valid": not errors, "errors": errors[:200], "valid_rows": len(rows) - len({item["row"] for item in errors})}


def commit_rows(db: Session, job: ContractImportJob, user_id: str) -> list[Contract]:
    rows = json.loads(job.rows_json or "[]")
    contracts: list[Contract] = []
    seen_numbers: set[str] = set()
    for index, row in enumerate(rows, start=2):
        parsed, errors = normalize_contract_row(row, index)
        if errors:
            raise HTTPException(status_code=409, detail="导入校验已失效，请重新校验")
        contract_no = parsed.get("contract_no")
        if contract_no:
            if contract_no in seen_numbers or db.query(Contract).filter(
                Contract.organization_id == job.organization_id, Contract.contract_no == contract_no
            ).first():
                raise HTTPException(status_code=409, detail="合同编号已存在，导入未写入任何合同")
            seen_numbers.add(contract_no)
        contracts.append(
            Contract(
                organization_id=job.organization_id,
                created_by=user_id,
                updated_by=user_id,
                source="import",
                **parsed,
            )
        )
    db.add_all(contracts)
    db.flush()
    return contracts
