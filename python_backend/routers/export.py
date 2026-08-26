"""Export analysis results to Excel."""

import io
import json

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_db
from models.document import AnalysisResult, Document

router = APIRouter(prefix="/api/export", tags=["export"])

HEADER_FONT = Font(bold=True, size=12)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=12, color="FFFFFF")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def parse_json(text: str) -> dict | None:
    """Extract JSON from response text."""
    if not text:
        return None
    import re

    try:
        m = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
        return json.loads((m.group(1) if m else text).strip())
    except (json.JSONDecodeError, AttributeError):
        return None


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_cell(ws, row, col, wrap=False):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    return cell


def excel_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def latest_result(db: Session, doc_id: str, prompt_type: str):
    return db.query(AnalysisResult).filter(
        AnalysisResult.document_id == doc_id,
        AnalysisResult.prompt_type == prompt_type,
    ).order_by(AnalysisResult.created_at.desc()).first()


def snapshot_fields(result) -> list[dict]:
    if not result or not result.fields_snapshot_json:
        return []
    try:
        value = json.loads(result.fields_snapshot_json)
        return value if isinstance(value, list) else []
    except json.JSONDecodeError:
        return []


@router.get("/{doc_id}")
def export_excel(doc_id: str, db: Session = Depends(get_db)):
    """Generate Excel file for a document's analysis results."""
    document = db.query(Document).filter(Document.id == doc_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="文档不存在")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "合同属性"
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 60
    ws1.merge_cells("A1:B1")
    attribute_record = latest_result(db, document.id, "attribute_extraction")
    title_suffix = (
        f" · {attribute_record.template_name} v{attribute_record.template_version}"
        if attribute_record and attribute_record.template_name
        else ""
    )
    ws1.cell(
        row=1,
        column=1,
        value=f"合同属性提取 — {document.original_filename}{title_suffix}",
    ).font = HEADER_FONT
    ws1.cell(row=2, column=1, value="属性").font = HEADER_FONT_WHITE
    ws1.cell(row=2, column=2, value="内容").font = HEADER_FONT_WHITE
    ws1.cell(row=2, column=1).fill = HEADER_FILL
    ws1.cell(row=2, column=2).fill = HEADER_FILL
    style_header(ws1, 2, 2)

    attr_result = parse_json(attribute_record.response_text) if attribute_record else None
    fields = snapshot_fields(attribute_record)
    row = 3
    if fields:
        attr_rows = [
            (field.get("label") or field.get("key"), attr_result.get(field.get("key"), "未提及"))
            for field in fields
        ] if attr_result else [
            (field.get("label") or field.get("key"), "（无分析结果）") for field in fields
        ]
    else:
        attr_rows = list(attr_result.items()) if attr_result else [("分析结果", "（无分析结果）")]
    for label, value in attr_rows:
        style_cell(ws1, row, 1).value = label
        style_cell(ws1, row, 2, wrap=True).value = excel_value(value)
        row += 1

    ws2 = wb.create_sheet("数据问题审查")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 40
    review_record = latest_result(db, document.id, "reasonability_check")
    review_result = parse_json(review_record.response_text) if review_record else None
    ws2.merge_cells("A1:F1")
    ws2.cell(row=1, column=1, value="数据问题审查").font = HEADER_FONT
    headers2 = ["项目", "合同标注", "描述/公式", "计算结果", "严重程度", "审查说明"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header(ws2, 2, 6)

    RED_FILL = PatternFill(start_color="F56C6C", end_color="F56C6C", fill_type="solid")
    RED_FONT = Font(color="FFFFFF", bold=True)
    YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    row = 3
    has_issue = False
    sev_order = {"严重": 0, "警告": 1, "注意": 2}
    items = review_result.get("数据问题", []) if review_result else []
    items.sort(key=lambda x: sev_order.get(x.get("严重程度", ""), 99))
    for item in items:
        style_cell(ws2, row, 1).value = item.get("项目", "")
        style_cell(ws2, row, 2).value = item.get("合同标注", "")
        style_cell(ws2, row, 3).value = item.get("验算公式", item.get("描述/公式", ""))
        style_cell(ws2, row, 4).value = item.get("验算结果", item.get("计算结果", ""))
        sev = item.get("严重程度", "")
        style_cell(ws2, row, 5).value = sev
        style_cell(ws2, row, 6, wrap=True).value = item.get("说明", "")
        if item.get("是否有问题") == "是":
            has_issue = True
            if sev == "严重":
                for c in range(1, 7):
                    ws2.cell(row=row, column=c).fill = RED_FILL
                    ws2.cell(row=row, column=c).font = RED_FONT
            else:
                for c in range(1, 7):
                    ws2.cell(row=row, column=c).fill = YELLOW_FILL
        row += 1
    if not has_issue:
        ws2.merge_cells(f"A{row}:F{row}")
        style_cell(ws2, row, 1).value = "各项数据暂未发现明显问题"
        ws2.cell(row=row, column=1).fill = OK_FILL

    ws3 = wb.create_sheet("内容合理性审查")
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 50
    ws3.merge_cells("A1:C1")
    ws3.cell(row=1, column=1, value="内容合理性审查").font = HEADER_FONT
    headers3 = ["方面", "评价", "建议"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header(ws3, 2, 3)
    row = 3
    if review_result:
        for item in review_result.get("内容合理性", []):
            style_cell(ws3, row, 1).value = item.get("方面", "")
            style_cell(ws3, row, 2, wrap=True).value = item.get("评价", "")
            style_cell(ws3, row, 3, wrap=True).value = item.get("建议", "")
            row += 1
        if review_result.get("总结"):
            row += 1
            ws3.merge_cells(f"A{row}:C{row}")
            style_cell(ws3, row, 1).value = "总结：" + review_result["总结"]
            ws3.cell(row=row, column=1).fill = WARN_FILL

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{document.original_filename.rsplit('.', 1)[0]}_分析结果.xlsx"
    from urllib.parse import quote as url_quote

    encoded_filename = url_quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )
